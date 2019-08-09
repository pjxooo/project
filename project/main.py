#!/usr/bin/env python
# -*- coding:utf-8 -*-
import linecache
from openpyxl import *
import re
import time
import pymysql
import random
from bs4 import BeautifulSoup
import threading
from selenium import webdriver

from bean.Abnormaloperation import Abnormaloperation
from bean.AdministrativeLicense import AdministrativeLicense
from bean.Administrativepunishment import Administrativepunishment
from bean.Branchstructure import Branchstructure
from bean.Brand import Brand
from bean.Changerecord import Changerecord
from bean.Company import Company
from bean.Initiatorscontributions import Initiatorscontributions
from bean.Investment import Investment
from bean.Judgment import Judgment
from bean.Keypersonnel import Keypersonnel
from bean.Patent import Patent
from bean.Qualitysupervisi import Qualitysupervisi
from bean.Shareholder import Shareholder
from bean.Softwarecopyright import Softwarecopyright
from bean.Trademark import Trademark
from bean.Websitefiling import Websitefiling
# 数据库连接
conn = pymysql.connect('192.168.2.253', 'test', '123456')
conn.select_db('cloud_account_system')
# 获取游标
cur = conn.cursor()
#爬取网址
url = 'https://xin.baidu.com'
def remove_duplicate(string): #字符去重
    ids = list(string.text)
    news_ids = []
    for id in ids:
        if id not in news_ids:
            news_ids.append(id)
    str2 = ''
    removename = str2.join(news_ids)
    return removename
def changepage(driver): #切换页面
    rul = driver.current_url # 获取新标签url
    driver.get(rul)
    driver.forward()
    return driver
def randomip(): #获取随机ip
    count = len(open('./ip.txt', 'rU').readlines())
    #生成随机行数
    hellonum = random.randrange(1, count, 1)
    #随机读取某行
    ranproxy = linecache.getline('./ip.txt', hellonum)
    return ranproxy
def dlip(): #使用代理ip打开网页
    chrome_options = webdriver.ChromeOptions()
    # chrome_options.add_argument("--proxy-server=http://183.165.31.215:23942")
    driver = webdriver.Chrome(chrome_options=chrome_options)
    driver.get(url)
    time.sleep(3)
    ifdriver = driver.page_source #判断是否成功连接代理ip
    ifdriveris1 = re.findall(r"无法访问此网站", ifdriver)
    ifdriveris2 = re.findall(r"未连接到互联网", ifdriver)
    if len(ifdriveris1) != 0 or len(ifdriveris2) !=0:
        print("代理ip连接失败")
        driver.quit()
    else:
        return driver
def company_detailinfo(driver): # 企业基本信息
    company = Company()
    company.setCompanyId(random.randint(10000000, 100000000))
    time.sleep(3)
    cname = driver.find_element_by_xpath(".//*[@class='zx-detail-company'][1]/h2/span").text
    company.setCompanyName(str(cname))
    cweb = driver.find_element_by_xpath(".//*[@class='zx-detail-company'][1]/div[2]/div[5]")
    ccweb = ''.join(re.findall(r"[^官网：]", cweb.text))
    company.setCompanyWebsite(str(ccweb))
    cphone = driver.find_element_by_xpath(".//*[@class='zx-detail-company'][1]/div[2]/div[3]")
    ccphone = ''.join(re.findall(r"[^电话：]", cphone.text))
    company.setCompanyPhone(str(ccphone))
    cemail = driver.find_element_by_xpath(".//*[@class='zx-detail-company'][1]/div[2]/div[4]")
    ccemail = ''.join(re.findall(r"[^邮箱：]", cemail.text))
    company.setCompanyEmail(str(ccemail))
    caddress = driver.find_element_by_xpath(".//*[@class='zx-detail-company'][1]/div[2]/div[6]")
    ccaddress = ''.join(re.findall(r"[^地址：]", caddress.text))
    company.setCompanyAddress(str(ccaddress))
    cinfo = driver.find_element_by_xpath(".//*[@class='zx-detail-company'][1]/div[2]/div[7]/p/span[1]")
    company.setCompanyInfo(str(cinfo.text))
    tds = driver.find_elements_by_xpath(".//*[@class='zx-detail-basic-table']/tbody/tr/td")
    capital = tds[1].text
    ccapital = ''.join(re.findall(r"[0-9]", capital))
    company.setRegisteredCapital(str(ccapital))
    company.setContributedMapital(str(tds[3].text))
    company.setLegalRepresentative(str(tds[5].text))
    company.setManagementForms(str(tds[7].text))
    company.setUsedName(str(tds[9].text))
    company.setIndustryInvolved(str(tds[11].text))
    company.setUnifiedSocialCreditCode(str(tds[13].text))
    company.setTaxpayerRegistrationNumber(str(tds[15].text))
    company.setBusinessRegistrationNumber(str(tds[17].text))
    company.setOrganizingInstitutionBarCode(str(tds[19].text))
    company.setRegistrationAuthority(str(tds[21].text))
    company.setDateOfEstablishment(str(tds[23].text))
    company.setTypeOfEnterprise(str(tds[25].text))
    company.setBusinessTerm(str(tds[27].text))
    company.setAdministrativeDivision(str(tds[29].text))
    company.setAnnualInspectionDate(str(tds[31].text))
    ra = tds[33].text
    rea = ''.join(re.findall(r"(.*)查看地图", ra))
    company.setRegisteredAddress(str(rea))
    company.setBusinessScope(str(tds[35].text))
    time.sleep(3)
    sql_company_insert = "insert into tbl_sycs_python_company(companyId,companyName,companyWebsite,companyPhone,companyEmail,companyAddress,companyInfo,registeredCapital,contributedMapital,legalRepresentative,managementForms,usedName,industryInvolved,unifiedSocialCreditCode,taxpayerRegistrationNumber,businessRegistrationNumber,organizingInstitutionBarCode,registrationAuthority,dateOfEstablishment,typeOfEnterprise,businessTerm,administrativeDivision,annualInspectionDate,registeredAddress,businessScope) values (%(companyId)s, %(companyName)s, %(companyWebsite)s, %(companyPhone)s, %(companyEmail)s, %(companyAddress)s, %(companyInfo)s, %(registeredCapital)s, %(contributedMapital)s,%(legalRepresentative)s,%(managementForms)s,%(usedName)s,%(industryInvolved)s,%(unifiedSocialCreditCode)s,%(taxpayerRegistrationNumber)s,%(businessRegistrationNumber)s,%(organizingInstitutionBarCode)s,%(registrationAuthority)s,%(dateOfEstablishment)s,%(typeOfEnterprise)s,%(businessTerm)s,%(administrativeDivision)s,%(annualInspectionDate)s,%(registeredAddress)s,%(businessScope)s)"
    company_message = {"companyId": company.getCompanyId(),
                       "companyName": company.getCompanyName(),
                       "companyWebsite": company.getCompanyWebsite(),
                       "companyPhone": company.getCompanyPhone(),
                       "companyEmail": company.getCompanyEmail(),
                       "companyAddress": company.getCompanyAddress(),
                       "companyInfo": company.getCompanyInfo(),
                       "registeredCapital": company.getRegisteredCapital(),
                       "contributedMapital": company.getContributedMapital(),
                       "legalRepresentative": company.getLegalRepresentative(),
                       "managementForms": company.getManagementForms(),
                       "usedName": company.getUsedName(),
                       "industryInvolved": company.getIndustryInvolved(),
                       "unifiedSocialCreditCode": company.getUnifiedSocialCreditCode(),
                       "taxpayerRegistrationNumber": company.getTaxpayerRegistrationNumber(),
                       "businessRegistrationNumber": company.getBusinessRegistrationNumber(),
                       "organizingInstitutionBarCode": company.getOrganizingInstitutionBarCode(),
                       "registrationAuthority": company.getRegistrationAuthority(),
                       "dateOfEstablishment": company.getDateOfEstablishment(),
                       "typeOfEnterprise": company.getTypeOfEnterprise(),
                       "businessTerm": company.getBusinessTerm(),
                       "administrativeDivision": company.getAdministrativeDivision(),
                       "annualInspectionDate": company.getAnnualInspectionDate(),
                       "registeredAddress": company.getRegisteredAddress(),
                       "businessScope": company.getBusinessScope()
                       }
    cur.execute(sql_company_insert, company_message)
    # conn.commit()
    return company
def company_detailinfo_shareholder(driver,company,companydetailinfo,soup):# 股东信息
    shareholer = Shareholder()  # 股东对象
    shareholerList = list()
    # 判断是否有股东
    ifhaveshareshoulder = re.findall(r"股东信息", companydetailinfo)
    if len(ifhaveshareshoulder) == 0:
        print("该公司没有股东信息" + "————————————" + str(company.getCompanyName()))
    else:
        c = soup.find("div", {"class": "basic-shareholder-container"}).table.tbody
        time.sleep(5)
        a = re.findall(r"<[a]", str(c))
        sp = re.findall(r"<[s]", str(c))
        la = len(a)
        lp = len(sp)
        table = driver.find_element_by_xpath(".//*[@class='basic-shareholder-container']/table/tbody")
        table_rows = table.find_elements_by_tag_name('tr')
        for sh in range(len(table_rows)):
            time.sleep(5)
            s = driver.find_elements_by_xpath(".//*[@class='basic-shareholder-container']/table/tbody/tr[%s]/td" % (sh + 1))
            if la == 0 and lp > 0:
                sname = driver.find_element_by_xpath(".//*[@class='basic-shareholder-container']/table/tbody/tr[%s]/td[2]/span" % (sh + 1))
                shareholer.setShareHolder(str(sname.text))
            elif la > 0 and lp == 0:
                sname = driver.find_element_by_xpath(".//*[@class='basic-shareholder-container']/table/tbody/tr[%s]/td[2]/a[1]" % (sh + 1))
                shareholer.setShareHolder(str(sname.text))
            elif la > 0 and lp > 0:
                sname = driver.find_element_by_xpath(".//*[@class='basic-shareholder-container']/table/tbody/tr[%s]/td[2]" % (sh + 1))
                _sname = remove_duplicate(sname)
                shareholer.setShareHolder(str(_sname))
            time.sleep(3)
            shareholer.setShareholdingRatio(str(s[2].text))
            shareholer.setSubscribedCapitalContribution(str(s[3].text))
            shareholer.setActualCapitalContribution(str(s[4].text))
            shareholer.setCompanyId(company.getCompanyId())
            shareholerList.append(shareholer)
            sql_shareholder_insert = "insert into tbl_sycs_python_shareholder(shareHolder,shareholdingRatio,subscribedCapitalContribution,actualCapitalContribution,companyId) values (%(shareHolder)s, %(shareholdingRatio)s, %(subscribedCapitalContribution)s, %(actualCapitalContribution)s, %(companyId)s)"
            shareholer_message = {"shareHolder": shareholerList[sh].getShareHolder(),
                                  "shareholdingRatio": shareholerList[sh].getShareholdingRatio(),
                                  "subscribedCapitalContribution": shareholerList[sh].getSubscribedCapitalContribution(),
                                  "actualCapitalContribution": shareholerList[sh].getActualCapitalContribution(),
                                  "companyId": shareholerList[sh].getCompanyId()}
            cur.execute(sql_shareholder_insert, shareholer_message)
            conn.commit()
def company_detailinfo_keypersonnel(driver,companydetailinfo,company):# 主要人员
    keypersonnel = Keypersonnel()  # 主要人员对象
    keypersonnelList = list()
    # 判断是否有主要人员
    ifhavekeypeisonnel = re.findall(r"职务", companydetailinfo)
    if len(ifhavekeypeisonnel) == 0:
        print("该公司没有主要人员信息" + "————————————" + str(company.getCompanyName()))
    else:
        time.sleep(5)
        kptable = driver.find_element_by_xpath(".//*[@class='basic-directors-container']/table/tbody")
        kptable_rows = kptable.find_elements_by_tag_name('tr')
        for k in range(len(kptable_rows)):
            kp = driver.find_elements_by_xpath(".//*[@class='basic-directors-container']/table/tbody/tr[%s]/td" % (k + 1))
            kpname = driver.find_element_by_xpath(".//*[@class='basic-directors-container']/table/tbody/tr[%s]/td[2]" % (k + 1))
            _kpname = remove_duplicate(kpname)
            keypersonnel.setKeyPersonnelName(str(_kpname))
            keypersonnel.setKeyPersonnelDuty(str(kp[2].text))
            keypersonnel.setCompanyId(company.getCompanyId())
            keypersonnelList.append(keypersonnel)
            sql_keypersonnel_insert = "insert into tbl_sycs_python_keypersonnel(keyPersonnelName,keyPersonnelDuty,companyId) values (%(keyPersonnelName)s, %(keyPersonnelDuty)s, %(companyId)s)"
            keypersonnel_message = {"keyPersonnelName": keypersonnelList[k].getKeyPersonnelName(),
                                    "keyPersonnelDuty": keypersonnelList[k].getKeyPersonnelDuty(),
                                    "companyId": keypersonnelList[k].getCompanyId()}
            cur.execute(sql_keypersonnel_insert, keypersonnel_message)
            conn.commit()
def company_detailinfo_invest(driver,companydetailinfo,company):# 对外投资
    investment = Investment()
    investmentList = list()
    def company_detailinfo_invest_rows():
        time.sleep(5)
        investtable = driver.find_element_by_xpath(".//*[@class='basic-invest-container']/table/tbody")
        investtable_rows = investtable.find_elements_by_tag_name('tr')
        for i in range(len(investtable_rows)):
            invest = driver.find_elements_by_xpath(".//*[@class='basic-invest-container']/table/tbody/tr[%s]/td" % (i + 1))
            investment.setInvestedEnterprise(str(driver.find_element_by_xpath(".//*[@class='basic-invest-container']/table/tbody/tr[%s]/td[2]/a" % (i + 1)).text))
            investment.setInvestedEnterpriseRepresentative(str(driver.find_element_by_xpath(".//*[@class='basic-invest-container']/table/tbody/tr[%s]/td[3]/span" % (i + 1)).text))
            investment.setEstablishmentDate(str(invest[3].text))
            investment.setInvestmentProportion(str(invest[4].text))
            investment.setSubscribedAmount(str(invest[5].text))
            investment.setState(str(invest[6].text))
            investment.setCompanyId(company.getCompanyId())
            investmentList.append(investment)
            sql_investment_insert = "insert into tbl_sycs_python_investment(investedEnterprise,investedEnterpriseRepresentative,establishmentDate,investmentProportion,subscribedAmount,state,companyId) values " \
                                    "(%(investedEnterprise)s, %(investedEnterpriseRepresentative)s, %(establishmentDate)s, %(investmentProportion)s, %(subscribedAmount)s, %(state)s, %(companyId)s)"
            investment_message = {"investedEnterprise": investmentList[i].getInvestedEnterprise(),
                                  "investedEnterpriseRepresentative": investmentList[i].getInvestedEnterpriseRepresentative(),
                                  "establishmentDate": investmentList[i].getEstablishmentDate(),
                                  "investmentProportion": investmentList[i].getInvestmentProportion(),
                                  "subscribedAmount": investmentList[i].getSubscribedAmount(),
                                  "state": investmentList[i].getState(),
                                  "companyId": investmentList[i].getCompanyId()}
            cur.execute(sql_investment_insert, investment_message)
            conn.commit()
    # 判断是否有对外投资
    ifhaveinvestment = re.findall(r"投资占比", companydetailinfo)
    if len(ifhaveinvestment) == 0:
        print("该公司没有对外投资" + "————————————" + str(company.getCompanyName()))
    else:
        # 判断投资企业是否超出10个，分页
        invest_rows = driver.find_element_by_xpath(".//*[@class='basic-invest-container']/h3/span")
        print(int(invest_rows.text))
        if int(invest_rows) > 10:
            company_detailinfo_invest_rows()
            a = driver.find_element_by_xpath(".//*[@class='basic-invest-pager ui-pager skin-cs skin-cs-pager']")
            a_rows = a.find_elements_by_tag_name('a')
            print(len(a_rows))
            for i in range(1, len(a_rows)-2):
                time.sleep(2)
                driver.find_element_by_xpath(".//*[@class='basic-invest-pager ui-pager skin-cs skin-cs-pager']/a[%s]" % len(a_rows)).click()
                company_detailinfo_invest_rows()
        else:
            company_detailinfo_invest_rows()
def company_detailinfo_changerecord(driver,companydetailinfo,company): #变更记录
    changerecord = Changerecord() #变更记录对象
    changerecordList = list()
    def company_detailinfo_changerecord_rows():
        time.sleep(5)
        crtable = driver.find_element_by_xpath(".//*[@class='basic-change-container']/table/tbody")
        crtable_rows = crtable.find_elements_by_tag_name('tr')
        for c in range(len(crtable_rows)):
            cr = driver.find_elements_by_xpath(".//*[@class='basic-change-container']/table/tbody/tr[%s]/td" % (c + 1))
            changerecord.setchangeRecordDate(str(cr[1].text))
            changerecord.setchangedItem(str(cr[2].text))
            changerecord.setbeforeChange(str(cr[3].text))
            aftercr = driver.find_elements_by_xpath(".//*[@class='basic-change-container']/table/tbody/tr[%s]/td[5]/span"% (c + 1))
            aftercr1 = []
            for m in aftercr:
                aftercr1.append(m.text)
            changerecord.setafterChange((" ".join('%s' % a for a in aftercr1)))
            changerecord.setCompanyId(company.getCompanyId())
            changerecordList.append(changerecord)
            sql_changerecord_insert = "insert into tbl_sycs_python_changerecord(changeRecordDate,changedItem,beforeChange,afterChange,companyId) values (%(changeRecordDate)s, %(changedItem)s, %(beforeChange)s, %(afterChange)s, %(companyId)s)"
            changerecord_message = {"changeRecordDate": changerecordList[c].getchangeRecordDate(),
                                    "changedItem": changerecordList[c].getchangedItem(),
                                    "beforeChange": changerecordList[c].getbeforeChange(),
                                    "afterChange": changerecordList[c].getafterChange(),
                                    "companyId": changerecordList[c].getCompanyId()}
            cur.execute(sql_changerecord_insert, changerecord_message)
            conn.commit()
    # 判断是否有变更记录
    ifhavechangerecord = re.findall(r"变更前", companydetailinfo)
    if len(ifhavechangerecord) == 0:
        print("该公司没有变更记录" + "————————————" + str(company.getCompanyName()))
    else:
        # 判断变更记录是否超出5个，分页
        record_rows = driver.find_element_by_xpath(".//*[@class='basic-change-container']/h3/span").text
        if int(record_rows) > 5:
            company_detailinfo_changerecord_rows()
            a = driver.find_element_by_xpath(".//*[@class='basic-change-pager ui-pager skin-cs skin-cs-pager']")
            a_rows = a.find_elements_by_tag_name('a')
            for i in range(1, len(a_rows) - 2):
                time.sleep(2)
                driver.find_element_by_xpath(".//*[@class='basic-change-pager ui-pager skin-cs skin-cs-pager']/a[%s]" % len(a_rows)).click()
                company_detailinfo_changerecord_rows()
        else:
            company_detailinfo_changerecord_rows()
def company_detailinfo_branchstructure(driver,companydetailinfo,company):  # 分支结构
    branchstructure = Branchstructure()
    branchstructureList = list()
    ifhavebranchstructure = re.findall(r"basic-branch-container", companydetailinfo)
    if len(ifhavebranchstructure) == 0:
        print("该公司没有分支机构" + "————————————" + str(company.getCompanyName()))
    else:
        time.sleep(5)
        branchstructuretable = driver.find_element_by_xpath(".//*[@class='basic-branch-container']/table/tbody")
        branchstructuretable_rows = branchstructuretable.find_elements_by_tag_name('tr')
        for i in range(len(branchstructuretable_rows)):
            structure = driver.find_elements_by_xpath(".//*[@class='basic-branch-container']/table/tbody/tr[%s]/td" % (i + 1))
            branchstructure.setCompanyName(str(driver.find_element_by_xpath(".//*[@class='basic-branch-container']/table/tbody/tr[%s]/td[2]/a" % (i + 1)).text))
            leader = driver.find_element_by_xpath(".//*[@class='basic-branch-container']/table/tbody/tr[%s]/td[3]" % (i + 1))
            _leader = remove_duplicate(leader)
            branchstructure.setLeader(_leader)
            branchstructure.setEstablishmentDate(str(structure[3].text))
            branchstructure.setState(str(structure[4].text))
            branchstructure.setCompanyId(company.getCompanyId())
            branchstructureList.append(branchstructure)
            sql_branchstructure_insert = "insert into tbl_sycs_python_branchstructure(companyName,leader,establishmentDate,state,companyId) values (%(companyName)s, %(leader)s, %(establishmentDate)s, %(state)s, %(companyId)s)"
            branchstructure_message = {"companyName": branchstructureList[i].getCompanyName(),
                                        "leader": branchstructureList[i].getLeader(),
                                        "establishmentDate": branchstructureList[i].getEstablishmentDate(),
                                        "state": branchstructureList[i].getState(),
                                        "companyId": branchstructureList[i].getCompanyId()}
            cur.execute(sql_branchstructure_insert, branchstructure_message)
            conn.commit()
def company_riskinfo_judgment(driver,company,companyriskinfo):  # 风险提示-裁判
    judgment = Judgment()  # 裁判文书对象
    judgmentList = list()
    def company_riskinfo_juagment_rows():
        time.sleep(5)
        judgmenttable = driver.find_element_by_xpath(".//*[@class='lawWenshu-table']/tbody")
        judgmenttable_rows = judgmenttable.find_elements_by_tag_name('tr')
        for i in range(1, len(judgmenttable_rows)):
            j = driver.find_elements_by_xpath(".//*[@class='lawWenshu-table']/tbody/tr[%s]/td" % (i + 1))
            judgment.setCaseDate(str(j[1].text))
            judgment.setCaseName(str(j[2].text))
            judgment.setCauseOfAction(str(j[3].text))
            judgment.setCaseStatus(str(j[4].text))
            judgment.setCaseNum(str(j[5].text))
            judgment.setCompanyId(company.getCompanyId())
            judgmentList.append(judgment)
            sql_judgment_insert = "insert into tbl_sycs_python_judgment(caseDate,caseName,causeOfAction,caseStatus,caseNum,companyId) values (%(caseDate)s, %(caseName)s, %(causeOfAction)s, %(caseStatus)s, %(caseNum)s, %(companyId)s)"
            judgment_message = {"caseDate": judgmentList[i - 1].getCaseDate(),
                                "caseName": judgmentList[i - 1].getCaseName(),
                                "causeOfAction": judgmentList[i - 1].getCauseOfAction(),
                                "caseStatus": judgmentList[i - 1].getCaseStatus(),
                                "caseNum": judgmentList[i - 1].getCaseNum(),
                                "companyId": judgmentList[i - 1].getCompanyId()}
            cur.execute(sql_judgment_insert, judgment_message)
            conn.commit()
    ifjudgment = re.findall(r"裁判文书", companyriskinfo)  # 判断是否有裁判文书
    if len(ifjudgment) == 0:
        print("该公司没有裁判文书")
    else:
        print("该公司有裁判文书")
        # 判断裁判文书是否超出5个，分页
        judgment_rows = driver.find_element_by_xpath(".//*[@class='zx-detail-lawWenshu-item']/h3/span").text
        if int(judgment_rows) > 5:
            company_riskinfo_juagment_rows()
            a = driver.find_element_by_xpath(".//*[@class='zx-detail-lawWenshu-pager ui-pager skin-cs skin-cs-pager']")
            a_rows = a.find_elements_by_tag_name('a')
            if len(a_rows)>7:
                for i in range(1, 6):
                    time.sleep(5)
                    driver.find_element_by_xpath(".//*[@class='zx-detail-lawWenshu-pager ui-pager skin-cs skin-cs-pager']/a[%s]" % len(a_rows)).click()
                    company_riskinfo_juagment_rows()
            else:
                for i in range(1, len(a_rows)-2):
                    time.sleep(5)
                    driver.find_element_by_xpath(".//*[@class='zx-detail-lawWenshu-pager ui-pager skin-cs skin-cs-pager']/a[%s]" % len(a_rows)).click()
                    company_riskinfo_juagment_rows()
        else:
            company_riskinfo_juagment_rows()
def company_riskinfo_punishment(driver, company, companyriskinfo):  # 风险提示-行政处罚
    punishment = Administrativepunishment()  # 行政处罚对象
    punishmentList = list()
    def company_riskinfo_punishment_rows():
        time.sleep(5)
        punishmenttable = driver.find_element_by_xpath(".//*[@class='penalties-table']/tbody")
        punishmenttable_rows = punishmenttable.find_elements_by_tag_name('tr')
        for i in range(1, len(punishmenttable_rows)):
            j = driver.find_elements_by_xpath(".//*[@class='penalties-table']/tbody/tr[%s]/td" % (i + 1))
            punishment.setDetermineInstrumentNumber(str(j[1].text))
            punishment.setAdministrativePunishmentType(str(j[2].text))
            punishment.setDecisiveOrganization(str(j[3].text))
            if str(j[4].text) == str("-"):
                punishment.setDecisiveDate(None)
            else:
                punishment.setDecisiveDate(str(j[4].text))
            punishment.setCompanyId(company.getCompanyId())
            punishmentList.append(punishment)
            sql_punishment_insert = "insert into tbl_sycs_python_administrativepunishment(determineInstrumentNumber,administrativePunishmentType,decisiveOrganization,decisiveDate,companyId) values (%(determineInstrumentNumber)s, %(administrativePunishmentType)s, %(decisiveOrganization)s, %(decisiveDate)s, %(companyId)s)"
            punishment_message = {"determineInstrumentNumber": punishmentList[i - 1].getDetermineInstrumentNumber(),
                                "administrativePunishmentType": punishmentList[i - 1].getAdministrativePunishmentType(),
                                "decisiveOrganization": punishmentList[i - 1].getDecisiveOrganization(),
                                "decisiveDate": punishmentList[i - 1].getDecisiveDate(),
                                "companyId": punishmentList[i - 1].getCompanyId()}
            cur.execute(sql_punishment_insert, punishment_message)
            conn.commit()
    ifpunishment = re.findall(r"行政处罚种类", companyriskinfo)  # 判断是否有裁判文书
    if len(ifpunishment) == 0:
        print("该公司没有行政处罚")
    else:
        print("该公司有行政处罚")
        # 判断行政处罚是否超出5个，分页
        punishment_rows = driver.find_element_by_xpath(".//*[@class='zx-detail-penalties-title']/span").text
        if int(punishment_rows) > 5:
            company_riskinfo_punishment_rows()
            a = driver.find_element_by_xpath(".//*[@class='zx-detail-penalties-pager ui-pager skin-cs skin-cs-pager']")
            a_rows = a.find_elements_by_tag_name('a')
            print(len(a_rows))
            for i in range(1, len(a_rows)-2):
                time.sleep(6)
                driver.find_element_by_xpath(".//*[@class='zx-detail-penalties-pager ui-pager skin-cs skin-cs-pager']/a[%s]" % len(a_rows)).click()
                company_riskinfo_punishment_rows()
        else:
            company_riskinfo_punishment_rows()
def company_riskinfo_abnormal(driver, company, companyriskinfo):  # 风险提示-经营异常
    abnormal = Abnormaloperation()
    abnormalList = list()
    ifhaveabnormal = re.findall(r"列入经营异常名录原因", companyriskinfo)
    if len(ifhaveabnormal) == 0:
        print("该公司没有经营异常" + "————————————" + str(company.getCompanyName()))
    else:
        time.sleep(5)
        abnormaltable = driver.find_element_by_xpath(".//*[@class='abnormal-table']/tbody")
        abnormaltable_rows = abnormaltable.find_elements_by_tag_name('tr')
        for i in range(1, len(abnormaltable_rows)):
            a = driver.find_elements_by_xpath(".//*[@class='abnormal-table']/tbody/tr[%s]/td" % (i + 1))
            abnormal.setInTime(str(a[1].text))
            abnormal.setInReason(str(a[2].text))
            abnormal.setInOrganization(str(a[3].text))
            if str(a[4].text) == str("-"):
                abnormal.setOutOfDate(None)
            else:
                abnormal.setOutOfDate(str(a[4].text))
            abnormal.setOutReason(str(a[5].text))
            abnormal.setOutOrganization(str(a[6].text))
            abnormal.setCompanyId(company.getCompanyId())
            abnormalList.append(abnormal)
            sql_abnormal_insert = "insert into tbl_sycs_python_abnormaloperation(inTime,inReason,inOrganization,outOfDate,outReason,outOrganization,companyId) values (%(inTime)s,%(inReason)s,%(inOrganization)s, %(outOfDate)s, %(outReason)s, %(outOrganization)s, %(companyId)s)"
            abnormal_message = {"inTime": abnormalList[i-1].getInTime(),
                                "inReason": abnormalList[i-1].getInReason(),
                                "inOrganization": abnormalList[i-1].getInOrganization(),
                                "outOfDate": abnormalList[i-1].getOutOfDate(),
                                "outReason": abnormalList[i - 1].getOutReason(),
                                "outOrganization": abnormalList[i - 1].getOutOrganization(),
                                "companyId": abnormalList[i-1].getCompanyId()}
            cur.execute(sql_abnormal_insert, abnormal_message)
            conn.commit()
def company_propertyinfo_websitefiling(driver, company, companypropertyinfo):  # 知识产权-网站备案
    website = Websitefiling()  # 网站备案对象
    websiteList = list()
    def company_propertyinfo_website_rows():
        time.sleep(5)
        websitetable = driver.find_element_by_xpath(".//*[@class='webRecord-table']/tbody")
        websitetable_rows = websitetable.find_elements_by_tag_name('tr')
        for i in range(len(websitetable_rows)):
            j = driver.find_elements_by_xpath(".//*[@class='webRecord-table']/tbody/tr[%s]/td" % (i + 1))
            website.setHomePage(str(j[1].text))
            website.setWebsiteName(str(j[2].text))
            dominname = driver.find_element_by_xpath(".//*[@class='webRecord-table']/tbody/tr[%s]/td/div/p" % (i + 1))
            website.setDomainName(str(dominname.text))
            website.setRecordNum(str(j[4].text))
            website.setCompanyId(company.getCompanyId())
            websiteList.append(website)
            sql_website_insert = "insert into tbl_sycs_python_websitefiling(homePage,websiteName,domainName,recordNum,companyId) values (%(homePage)s, %(websiteName)s, %(domainName)s, %(recordNum)s, %(companyId)s)"
            website_message = {"homePage": websiteList[i].getHomePage(),
                                "websiteName": websiteList[i].getWebsiteName(),
                                "domainName": websiteList[i].getDomainName(),
                                "recordNum": websiteList[i].getRecordNum(),
                                "companyId": websiteList[i].getCompanyId()}
            cur.execute(sql_website_insert, website_message)
            # conn.commit()
    ifwebsite = re.findall(r"网站名称", companypropertyinfo)  # 判断是否有网站备案
    if len(ifwebsite) == 0:
        print("该公司没有网站备案")
    else:
        print("该公司有网站备案")
        # 判断网站备案是否超出10个，若超过截取前三页信息
        website_rows = driver.find_element_by_xpath(".//*[@class='zx-detail-icp-title']/span").text
        if int(website_rows) > 10:
            company_propertyinfo_website_rows()
            a = driver.find_element_by_xpath(".//*[@class='zx-webRecord-pager ui-pager skin-cs skin-cs-pager']")
            a_rows = a.find_elements_by_tag_name('a')
            if len(a_rows)> 5:
                for i in range(1, 3):
                    driver.find_element_by_xpath(".//*[@class='zx-webRecord-pager ui-pager skin-cs skin-cs-pager']/a[%s]" % len(a_rows)).click()
                    time.sleep(5)
                    company_propertyinfo_website_rows()
            else:
                for i in range(1, len(a_rows) - 2):
                    driver.find_element_by_xpath(".//*[@class='zx-webRecord-pager ui-pager skin-cs skin-cs-pager']/a[%s]" % len(a_rows)).click()
                    time.sleep(5)
                    company_propertyinfo_website_rows()
        else:
            company_propertyinfo_website_rows()
def company_propertyinfo_trademark(driver, company, companypropertyinfo):  # 知识产权-商标信息
    trademark = Trademark()  # 商标对象
    trademarkList = list()
    def company_propertyinfo_trademark_rows():
        time.sleep(5)
        trademarktable = driver.find_element_by_xpath(".//*[@class='zx-detail-markinfo-list']")
        trademarktable_rows = trademarktable.find_elements_by_tag_name('li')
        for i in range(len(trademarktable_rows)):
            trademarkname = driver.find_elements_by_xpath(".//*[@class='zx-detail-markinfo-list']/li[%s]/div/div[2]" % (i + 1))
            _trademarkname = ''.join(re.findall(r"[^商标名称:]", trademarkname.text))
            trademark.setTrademarkName(str(_trademarkname))
            registnum = driver.find_elements_by_xpath(".//*[@class='zx-detail-markinfo-list']/li[%s]/div/div[3]/div[1]" % (i + 1))
            _registnum = ''.join(re.findall(r"[^注册号:]", registnum.text))
            trademark.setRegistrationNumber(str(_registnum))
            trademarktype = driver.find_elements_by_xpath(".//*[@class='zx-detail-markinfo-list']/li[%s]/div/div[4]/div[1]" % (i + 1))
            _trademarktype = ''.join(re.findall(r"[^商标类型:]", trademarktype.text))
            trademark.setTrademarkType(str(_trademarktype))
            effectivetime = driver.find_elements_by_xpath(".//*[@class='zx-detail-markinfo-list']/li[%s]/div/div[3]/div[2]" % (i + 1))
            _effectivetime = ''.join(re.findall(r"[^商标有效时间:]", effectivetime.text))
            trademark.setEffectiveTime(str(_effectivetime))
            processtatus = driver.find_elements_by_xpath(".//*[@class='zx-detail-markinfo-list']/li[%s]/div/div[4]/div[2]" % (i + 1))
            _processtatus = ''.join(re.findall(r"[^ 商标流程状态:]", processtatus.text))
            trademark.setProcessStatus(str(_processtatus))
            trademark.setCompanyId(company.getCompanyId())
            trademarkList.append(trademark)
            sql_trademark_insert = "insert into tbl_sycs_python_websitefiling(trademarkName,registrationNumber,trademarkType,effectiveTime,processStatus,companyId) values (%(trademarkName)s, %(registrationNumber)s, %(trademarkType)s, %(effectiveTime)s, %(processStatus)s, %(companyId)s)"
            trademark_message = {"trademarkName": trademarkList[i].getTrademarkName(),
                                "registrationNumber": trademarkList[i].getRegistrationNumber(),
                                "trademarkType": trademarkList[i].getTrademarkType(),
                                "effectiveTime": trademarkList[i].getEffectiveTime(),
                                "processStatus": trademarkList[i].getProcessStatus(),
                                "companyId": trademarkList[i].getCompanyId()}
            cur.execute(sql_trademark_insert, trademark_message)
            conn.commit()
    iftrademark = re.findall(r"商标名称", companypropertyinfo)  # 判断是否有商标信息
    if len(iftrademark) == 0:
        print("该公司没有商标信息")
    else:
        print("该公司有商标信息")
        # 判断商标信息是否超出10个，若超过截取前三页信息
        trademark_rows = driver.find_element_by_xpath(".//*[@class='zx-detail-markinfo-count-num']").text
        if int(trademark_rows) > 10:
            company_propertyinfo_trademark_rows()
            a = driver.find_element_by_xpath(".//*[@class='zx-mark-pager ui-pager skin-cs skin-cs-pager']")
            a_rows = a.find_elements_by_tag_name('a')
            if len(a_rows) > 5:
                for i in range(1, 4):
                    driver.find_element_by_xpath(".//*[@class='zx-mark-pager ui-pager skin-cs skin-cs-pager']/a[%s]" % len(a_rows)).click()
                    time.sleep(5)
                    company_propertyinfo_trademark_rows()
            else:
                for i in range(1, len(a_rows) - 2):
                    driver.find_element_by_xpath(".//*[@class='zx-mark-pager ui-pager skin-cs skin-cs-pager']/a[%s]" % len(a_rows)).click()
                    time.sleep(5)
                    company_propertyinfo_trademark_rows()
        else:
            company_propertyinfo_trademark_rows()
def company_propertyinfo_patent(driver, company, companypropertyinfo):  # 知识产权-专利信息
    patent = Patent()  # 专利对象
    patentList = list()
    def company_propertyinfo_patent_rows():
        time.sleep(5)
        patentable = driver.find_element_by_xpath(".//*[@class='zx-detail-patentinfo-list']")
        patentable_rows = patentable.find_elements_by_tag_name('div')
        for i in range(len(patentable_rows)):
            patentdiv = driver.find_element_by_xpath(".//*[@class='zx-detail-patentinfo-list']/div[%s]" % (i+1))
            patendiv_rows = patentdiv.find_elements_by_tag_name('li')
            for j in range(len(patendiv_rows)):
                patent.setPatentName(str(driver.find_element_by_xpath(".//*[@class='zx-detail-patentinfo-list']/div[%s]/li[%s]/div[3]/span" % (i+1, j+1)).text))
                application = driver.find_element_by_xpath(".//*[@class='zx-detail-patentinfo-list']/div[%s]/li[%s]/div[4]/div" % (i+1, j+1))
                _application = ''.join(re.findall(r"[^申请公布号:]", application.text))
                patent.setApplicationPublicationNum(str(_application))
                patentype = driver.find_element_by_xpath(".//*[@class='zx-detail-patentinfo-list']/div[%s]/li[%s]/div[5]/div" % (i+1, j+1))
                _pantentype = ''.join(re.findall(r"[^专利类型:]", patentype.text))
                patent.setPatentType(str(_pantentype))
                publicationtime = driver.find_element_by_xpath(".//*[@class='zx-detail-patentinfo-list']/div[%s]/li[%s]/div[6]/div" % (i+1, j+1))
                _publicationtime = ''.join(re.findall(r"[^公布日期:]", publicationtime.text))
                patent.setPublicationTime(str(_publicationtime))
                patent.setCompanyId(company.getCompanyId())
                patentList.append(patent)
                sql_trademark_insert = "insert into tbl_sycs_python_websitefiling(patentName,applicationPublicationNum,patentType,publicationTime,companyId) values (%(patentName)s, %(applicationPublicationNum)s, %(patentType)s, %(publicationTime)s, %(companyId)s)"
                trademark_message = {"patentName": patentList[i].getPatentName(),
                                     "applicationPublicationNum": patentList[i].getApplicationPublicationNum(),
                                     "patentType": patentList[i].getPatentType(),
                                     "publicationTime": patentList[i].getPublicationTime(),
                                     "companyId": patentList[i].getCompanyId()}
                cur.execute(sql_trademark_insert, trademark_message)
                conn.commit()
    ifpatent = re.findall(r"专利名称", companypropertyinfo)  # 判断是否有商标信息
    if len(ifpatent) == 0:
        print("该公司没有专利信息")
    else:
        print("该公司有专利信息")
        # 判断专利信息是否超出10个，若超过截取前三页信息，若没超过按照长度截取全部信息
        trademark_rows = driver.find_element_by_xpath(".//*[@class='zx-detail-patentinfo-count-num']").text
        if int(trademark_rows) > 10:
            company_propertyinfo_patent_rows()
            a = driver.find_element_by_xpath(".//*[@class='zx-patent-pager ui-pager skin-cs skin-cs-pager']")
            a_rows = a.find_elements_by_tag_name('a')
            if len(a_rows) > 5:
                for i in range(1, 4):
                    time.sleep(5)
                    driver.find_element_by_xpath(".//*[@class='zx-patent-pager ui-pager skin-cs skin-cs-pager']/a[%s]" % len(a_rows)).click()
                    company_propertyinfo_patent_rows()
            else:
                for i in range(1, len(a_rows) - 2):
                    time.sleep(5)
                    driver.find_element_by_xpath(".//*[@class='zx-patent-pager ui-pager skin-cs skin-cs-pager']/a[%s]" % len(a_rows)).click()
                    company_propertyinfo_patent_rows()
        else:
            company_propertyinfo_patent_rows()
def company_propertyinfo_brand(driver, company, companypropertyinfo): # 知识产权-品牌信息
    brand = Brand()  # 品牌对象
    brandList = list()
    def company_propertyinfo_brand_rows():
        time.sleep(5)
        brandtable = driver.find_element_by_xpath(".//*[@class='zx-detail-brandList-list']")
        brandtable_rows = brandtable.find_elements_by_tag_name('li')
        for i in range(len(brandtable_rows)):
            brandname = driver.find_element_by_xpath(".//*[@class='zx-detail-brandList-list']/li[%s]/a/div[2]/p[1]" % i + 1)
            _brandname = ''.join(re.findall(r"[^中文名称：]", brandname.text))
            brand.setBrandName(str(_brandname))
            origin = driver.find_element_by_xpath(".//*[@class='zx-detail-brandList-list']/li[%s]/a/div[2]/p[2]" % i + 1)
            _origin = ''.join(re.findall(r"[^发源地：]", origin.text))
            brand.setOrigin(str(_origin))
            year = driver.find_element_by_xpath(".//*[@class='zx-detail-brandList-list']/li[%s]/a/div[2]/p[3]" % i + 1)
            _year = ''.join(re.findall(r"[^创建年份：]", year.text))
            brand.setCreatedYear(str(_year))
            brand.setCompanyId(company.getCompanyId())
            brandList.append(brand)
            sql_brand_insert = "insert into tbl_sycs_python_brand(brandName,origin,createdYear,companyId) values (%(brandName)s, %(origin)s, %(createdYear)s, %(companyId)s)"
            brand_message = {"brandName": brandList[i].getBrandName(),
                            "origin": brandList[i].getOrigin(),
                            "createdYear": brandList[i].getCreatedYear(),
                            "companyId": brandList[i].getCompanyId()}
            cur.execute(sql_brand_insert, brand_message)
            conn.commit()
    ifpatent = re.findall(r"zx-brandList-content", companypropertyinfo)  # 判断是否有品牌信息
    if len(ifpatent) == 0:
        print("该公司没有品牌信息")
    else:
        print("该公司有品牌信息")
        # 判断品牌信息是否超出10个，若超过截取前三页信息，若没超过按照长度截取全部信息
        brand_rows = driver.find_element_by_xpath(".//*[@class='zx-detail-brandList-count-num']").text
        if int(brand_rows) > 10:
            company_propertyinfo_brand_rows()
            a = driver.find_element_by_xpath(".//*[@class='zx-brandList-pager ui-pager skin-cs skin-cs-pager']")
            a_rows = a.find_elements_by_tag_name('a')
            if len(a_rows) > 6:
                for i in range(1, 5):
                    time.sleep(5)
                    driver.find_element_by_xpath(".//*[@class='zx-brandList-pager ui-pager skin-cs skin-cs-pager']/a[%s]" % len(a_rows)).click()
                    company_propertyinfo_brand_rows()
            else:
                for i in range(1, len(a_rows) - 2):
                    time.sleep(5)
                    driver.find_element_by_xpath(".//*[@class='zx-brandList-pager ui-pager skin-cs skin-cs-pager']/a[%s]" % len(a_rows)).click()
                    company_propertyinfo_brand_rows()
        else:
            company_propertyinfo_brand_rows()
def company_propertyinfo_softwarecopyright(driver, company, companypropertyinfo):  #知识产权-软件著作信息
    software = Softwarecopyright()  # 软件著作对象
    softwareList = list()
    def company_propertyinfo_software_rows():
        time.sleep(5)
        softwaretable = driver.find_element_by_xpath(".//*[@class='zx-detail-copyrightinfo-list']")
        softwaretable_rows = softwaretable.find_elements_by_tag_name('div')
        for i in range(len(softwaretable_rows)):
            softwarediv = driver.find_element_by_xpath(".//*[@class='zx-detail-copyrightinfo-list']/div[%s]" % (i + 1))
            softwarediv_rows = softwarediv.find_elements_by_tag_name('li')
            for j in range(len(softwarediv_rows)):
                software.setSofewareWorksName(str(driver.find_element_by_xpath(".//*[@class='zx-detail-patentinfo-list']/div[%s]/li[%s]/div[3]/span" % (i + 1, j + 1)).text))
                versionnum = driver.find_element_by_xpath(".//*[@class='zx-detail-patentinfo-list']/div[%s]/li[%s]/div[4]/div" % (i + 1, j + 1))
                _versionnum = ''.join(re.findall(r"[^版本号：]", versionnum.text))
                software.setVersionNum(str(_versionnum))
                registdate = driver.find_element_by_xpath(".//*[@class='zx-detail-patentinfo-list']/div[%s]/li[%s]/div[5]/div" % (i + 1, j + 1))
                _registdate = ''.join(re.findall(r"[^登记日期]", registdate.text))
                software.setRegistrationDate(str(_registdate))
                software.setCompanyId(company.getCompanyId())
                softwareList.append(software)
                sql_software_insert = "insert into tbl_sycs_python_softwarecopyright(sofewareWorksName,versionNum,registrationDate,companyId) values (%(sofewareWorksName)s, %(versionNum)s, %(registrationDate)s, %(companyId)s)"
                software_message = {"sofewareWorksName": softwareList[i].getSofewareWorksName(),
                                     "versionNum": softwareList[i].getVersionNum(),
                                     "registrationDate": softwareList[i].getRegistrationDate(),
                                     "companyId": softwareList[i].getCompanyId()}
                cur.execute(sql_software_insert, software_message)
                conn.commit()
    ifsoftware = re.findall(r"版本号", companypropertyinfo)  # 判断是否有软件著作信息
    if len(ifsoftware) == 0:
        print("该公司没有软件著作信息")
    else:
        print("该公司有软件著作信息")
        # 判断软件著作信息是否超出10个，若超过截取前三页信息，若没超过按照长度截取全部信息
        software_rows = driver.find_element_by_xpath(".//*[@class='zx-detail-copyrightinfo-count-num']").text
        if int(software_rows) > 10:
            company_propertyinfo_software_rows()
            a = driver.find_element_by_xpath(".//*[@class='zx-copyright-pager ui-pager skin-cs skin-cs-pager']")
            a_rows = a.find_elements_by_tag_name('a')
            if len(a_rows) > 5:
                for i in range(1, 4):
                    time.sleep(5)
                    driver.find_element_by_xpath(".//*[@class='zx-copyright-pager ui-pager skin-cs skin-cs-pager']/a[%s]" % len(a_rows)).click()
                    company_propertyinfo_software_rows()
            else:
                for i in range(1, len(a_rows) - 2):
                    time.sleep(5)
                    driver.find_element_by_xpath(".//*[@class='zx-copyright-pager ui-pager skin-cs skin-cs-pager']/a[%s]" % len(a_rows)).click()
                    company_propertyinfo_software_rows()
        else:
            company_propertyinfo_software_rows()
def company_annualreportinfo_initiatorscontributions(driver, company, companyannualreportinfo):  # 年报信息-发起人及投资
    initiator = Initiatorscontributions()
    initiatorList = list()
    def company_annualreportinfo_initiatorscontributions_rows(year):
        ifhaveinitiator = re.findall(r"出资额", companyannualreportinfo)
        if len(ifhaveinitiator) != 0:  # 发起人及出资不为空
            initiatortable = driver.find_element_by_xpath(".//*[@class='annual-table table-center'][2]/tbody")
            initiatortable_rows = initiatortable.find_elements_by_tag_name('tr')
            for j in range(4, len(initiatortable_rows) + 1):
                initiator.setAnnualReportYear(year)
                tds = driver.find_elements_by_xpath(".//*[@class='annual-table table-center'][2]/tbody/tr[%s]/td" % j)
                initiator.setInitiator(str(tds[0].text))
                initiator.setCapitalContribution(str(tds[1].text))
                initiator.setContributionTime(str(tds[2].text))
                initiator.setModeOfInvestment(str(tds[3].text))
                initiator.setCompanyId(company.getCompanyId())
                initiatorList.append(initiator)
                sql_initiator_insert = "insert into tbl_sycs_python_initiatorscontributions(annualReportYear,initiator,capitalContribution,contributionTime,modeOfInvestment,companyId) values (%(annualReportYear)s, %(initiator)s, %(capitalContribution)s, %(contributionTime)s, %(modeOfInvestment)s, %(companyId)s)"
                initiator_message = {"annualReportYear": initiatorList[j - 4].getAnnualReportYear(),
                                    "initiator": initiatorList[j - 4].getInitiator(),
                                     "capitalContribution": initiatorList[j - 4].getCapitalContribution(),
                                     "contributionTime": initiatorList[j - 4].getContributionTime(),
                                     "modeOfInvestment": initiatorList[j - 4].getModeOfInvestment(),
                                     "companyId": initiatorList[j - 4].getCompanyId()}
                cur.execute(sql_initiator_insert, initiator_message)
                conn.commit()
        else:
            print("发起人及投资：无")
    initiatordiv = driver.find_element_by_xpath(".//*[@class='zx-detail-annual-tab']")
    initiatordiv_rows = initiatordiv.find_elements_by_tag_name('div')
    if len(initiatordiv_rows) > 1:  # 有一年以上的企业年报
        company_annualreportinfo_initiatorscontributions_rows(str(driver.find_element_by_xpath(".//*[@class='zx-detail-annual-tab']/div[1]").text))
        for i in range(len(initiatordiv_rows)-1):
            driver.find_element_by_xpath(".//*[@class='zx-detail-annual-tab']/div[%s]" % (i+2)).click()
            time.sleep(3)
            year = driver.find_element_by_xpath(".//*[@class='zx-detail-annual-tab']/div[%s]" % (i+2)).text
            company_annualreportinfo_initiatorscontributions_rows(str(year))
            pass
    else:  # 只有一年的企业年报
        year = driver.find_element_by_xpath(".//*[@class='zx-detail-annual-tab']/div[1]")
        company_annualreportinfo_initiatorscontributions_rows(year)
def company_operationalstatus_administrativeLicense(driver, company, operational_status):  # 经营状况-行政许可
    aLicense = AdministrativeLicense()  # 行政许可对象
    aLicenseList = list()
    def company_operationalstatus_administrativeLicense_rows():
        time.sleep(5)
        aLicensetable = driver.find_element_by_xpath(".//*[@class='condition-license-container']/table/tbody")
        aLicensetable_rows = aLicensetable.find_elements_by_tag_name('tr')
        for i in range(len(aLicensetable_rows)):
            j = driver.find_elements_by_xpath(".//*[@class='condition-license-container']/table/tbody/tr[%s]/td" % (i + 1))
            aLicense.setAdministrativeLicenseNum(str(j[1].text))
            aLicense.setLicenseName(str(j[2].text))
            aLicense.setLicenseContent(str(j[3].text))
            if str(j[4].text) == str("-"):
                aLicense.setBeginOfValidity(None)
            else:
                aLicense.setBeginOfValidity(str(j[4].text))
            if str(j[5].text) == str("-"):
                aLicense.setStopOfValidity(None)
            else:
                aLicense.setStopOfValidity(str(j[5].text))
            aLicense.setLicensingAuthority(str(j[6].text))
            aLicense.setCompanyId(company.getCompanyId())
            aLicenseList.append(aLicense)
            sql_aLicense_insert = "insert into tbl_sycs_python_administrativeLicense(administrativeLicenseNum,licenseName,licenseContent,beginOfValidity,stopOfValidity,licensingAuthority,companyId) values (%(administrativeLicenseNum)s, %(licenseName)s, %(licenseContent)s, %(beginOfValidity)s, %(stopOfValidity)s, %(licensingAuthority)s, %(companyId)s)"
            aLicense_message = {"administrativeLicenseNum": aLicenseList[i].getAdministrativeLicenseNum(),
                                "licenseName": aLicenseList[i].getLicenseName(),
                                "licenseContent": aLicenseList[i].getLicenseContent(),
                                "beginOfValidity": aLicenseList[i].getBeginOfValidity(),
                                "stopOfValidity": aLicenseList[i].getStopOfValidity(),
                                "licensingAuthority": aLicenseList[i].getLicensingAuthority(),
                                "companyId": aLicenseList[i].getCompanyId()}
            cur.execute(sql_aLicense_insert, aLicense_message)
            conn.commit()

    ifaLicense = re.findall(r"行政许可证号", operational_status)  # 判断是否有行政许可
    if len(ifaLicense) == 0:
        print("该公司没有行政许可")
    else:
        print("该公司有行政许可")
        # 判断行政许可是否超出5个，分页
        aLicense_rows = driver.find_element_by_xpath(".//*[@class='condition-license-container']/h3/span").text
        if int(aLicense_rows) > 5:
            company_operationalstatus_administrativeLicense_rows()
            a = driver.find_element_by_xpath(".//*[@class='condition-license-pager ui-pager skin-cs skin-cs-pager']")
            a_rows = a.find_elements_by_tag_name('a')
            if len(a_rows) > 5:
                for i in range(1, 3):
                    driver.find_element_by_xpath(".//*[@class='condition-license-pager ui-pager skin-cs skin-cs-pager']/a[%s]" % len(a_rows)).click()
                    time.sleep(5)
                    company_operationalstatus_administrativeLicense_rows()
            else:
                for i in range(1, len(a_rows) - 2):
                    driver.find_element_by_xpath(".//*[@class='condition-license-pager ui-pager skin-cs skin-cs-pager']/a[%s]" % len(a_rows)).click()
                    time.sleep(5)
                    company_operationalstatus_administrativeLicense_rows()
        else:
            company_operationalstatus_administrativeLicense_rows()
def company_operationalstatus_qualitysupervisi(driver, company, operational_status):  # 经营状况-质量监督检查
    qsupervisi = Qualitysupervisi()  # 质量监督对象
    qsupervisiList = list()
    def company_operationalstatus_qsupervisi_rows():
        time.sleep(5)
        qsupervisitable = driver.find_element_by_xpath(".//*[@class='condition-quality-container']/table/tbody")
        qsupervisitable_rows = qsupervisitable.find_elements_by_tag_name('tr')
        for i in range(len(qsupervisitable_rows)):
            j = driver.find_elements_by_xpath(".//*[@class='condition-quality-container']/table/tbody/tr[%s]/td" % (i + 1))
            qsupervisi.setSampleYear(str(j[1].text))
            qsupervisi.setQualitySupervisiIdBatches(str(j[2].text))
            qsupervisi.setQualitySupervisiProducts(str(j[3].text))
            qsupervisi.setSamplingResults(str(j[4].text))
            qsupervisi.setCompanyId(company.getCompanyId())
            qsupervisiList.append(qsupervisi)
            sql_qsupervisi_insert = "insert into tbl_sycs_python_qualitysupervisi(sampleYear,qualitySupervisiIdBatches,qualitySupervisiProducts,samplingResults,companyId) values (%(sampleYear)s, %(qualitySupervisiIdBatches)s, %(qualitySupervisiProducts)s, %(samplingResults)s, %(companyId)s)"
            qsupervisi_message = {"sampleYear": qsupervisiList[i].getSampleYear(),
                                    "qualitySupervisiIdBatches": qsupervisiList[i].getQualitySupervisiIdBatches(),
                                    "qualitySupervisiProducts": qsupervisiList[i].getQualitySupervisiProducts(),
                                    "samplingResults": qsupervisiList[i].getSamplingResults(),
                                    "companyId": qsupervisiList[i].getCompanyId()}
            cur.execute(sql_qsupervisi_insert, qsupervisi_message)
            conn.commit()
    ifqsupervisi = re.findall(r"质量监督抽查批次", operational_status)  # 判断是否有质量监督检查
    if len(ifqsupervisi) == 0:
        print("该公司没有质量监督检查")
    else:
        print("该公司有质量监督检查")
        # 判断质量监督检查是否超出5个，分页
        qsupervisi_rows = driver.find_element_by_xpath(".//*[@class='condition-quality-container']/h3/span").text
        if int(qsupervisi_rows) > 5:
            company_operationalstatus_qsupervisi_rows()
            a = driver.find_element_by_xpath(".//*[@class='condition-quality-pager ui-pager skin-cs skin-cs-pager']")
            a_rows = a.find_elements_by_tag_name('a')
            if len(a_rows) > 5:  # 如果大于5，截取前三页的数据
                for i in range(1, 3):
                    driver.find_element_by_xpath(".//*[@class='condition-quality-pager ui-pager skin-cs skin-cs-pager']/a[%s]" % len(a_rows)).click()
                    time.sleep(5)
                    company_operationalstatus_qsupervisi_rows()
            else:
                for i in range(1, len(a_rows) - 2):
                    driver.find_element_by_xpath(".//*[@class='condition-quality-pager ui-pager skin-cs skin-cs-pager']/a[%s]" % len(a_rows)).click()
                    time.sleep(5)
                    company_operationalstatus_qsupervisi_rows()
        else:
            company_operationalstatus_qsupervisi_rows()
def company_info(companyname):
    driver = dlip()
    time.sleep(2)
    if driver != None:
        # 输入企业名称并搜索
        driver.find_element_by_class_name("search-text").send_keys(companyname)
        driver.find_element_by_class_name("search-btn").click()
        time.sleep(2)
        # 切换到查询公司结果界面
        driver = changepage(driver)
        companies = driver.page_source
        # 判断是否有该公司
        ifhavecompany = re.findall(r"抱歉，没有找到相关结果...", companies)
        if len(ifhavecompany) != 0:
            print("没有找到该公司" + "———————————" + str(companyname))
            driver.quit()
        else:
            time.sleep(8)
            driver.find_elements_by_xpath(".//*[@class='zx-ent-logo']")[0].click()
            time.sleep(3)
            # 切换当前页面标签
            driver.switch_to.window(driver.window_handles[1])
            # 切换到企业基本信息界面
            driver = changepage(driver)
            companydetailinfo = driver.page_source
            soup = BeautifulSoup(companydetailinfo, features='lxml')
            ifloading = re.findall(r"无法访问此网站", companydetailinfo)
            if len(ifloading) != 0:
                print("详情页连接失败" + "————————————" + str(companyname))
                driver.quit()
            else:
                time.sleep(8)
                company = company_detailinfo(driver)
                time.sleep(5)
                # company_detailinfo_shareholder(driver, company, companydetailinfo, soup)
                # time.sleep(5)
                # company_detailinfo_keypersonnel(driver, companydetailinfo, company)
                # time.sleep(5)
                # company_detailinfo_invest(driver, companydetailinfo, company)
                # time.sleep(5)
                # company_detailinfo_changerecord(driver, companydetailinfo, company)
                # time.sleep(5)
                # company_detailinfo_branchstructure(driver, companydetailinfo, company)
                # 风险提示
                # risknum = driver.find_element_by_xpath(".//*[@class='table']/tbody/tr/td[2]/span[2]").text
                # if risknum == 0:
                #     print("该公司没有风险提示")
                # else:
                #     print("该公司有风险提示")
                #     driver.find_element_by_xpath(".//*[@class='table']/tbody/tr/td[2]").click()
                #     driver.switch_to.window(driver.window_handles[1])  # 切换当前页面标签
                #     driver = changepage(driver)  # 切换到风险提示页面
                #     companyriskinfo = driver.page_source
                #     company_riskinfo_judgment(driver, company, companyriskinfo)  # 裁判文书
                #     company_riskinfo_punishment(driver, company, companyriskinfo)  # 行政处罚
                #     company_riskinfo_abnormal(driver, company, companyriskinfo)  # 经营异常
                # 知识产权
                propertynum = driver.find_element_by_xpath(".//*[@class='table']/tbody/tr/td[3]/span[2]").text
                if propertynum == 0:
                    print("该公司没有知识产权")
                else:
                    print("该公司有知识产权")
                    driver.find_element_by_xpath(".//*[@class='table']/tbody/tr/td[3]").click()
                    driver.switch_to.window(driver.window_handles[1])  # 切换当前页面标签
                    driver = changepage(driver)  # 切换到知识产权页面
                    companypropertyinfo = driver.page_source
                    time.sleep(5)
                    # driver.find_element_by_xpath(".//*[@class='zx-detail-cert']/ul/li[1]").click()
                    time.sleep(5)
                    company_propertyinfo_websitefiling(driver, company, companypropertyinfo)  # 网站备案
                    time.sleep(10)
                    driver.find_element_by_xpath(".//*[@class='zx-detail-cert']']/ul/li[2]").click()
                    time.sleep(5)
                    company_propertyinfo_brand(driver, company, companypropertyinfo)  # 品牌信息


                # 企业年报-发起人及投资
                # annualreportnum = driver.find_element_by_xpath(".//*[@class='table']/tbody/tr/td[4]/span[2]").text
                # if annualreportnum == 0:
                #     print("该公司没有年报信息")
                # else:
                #     print("该公司有年报信息")
                #     driver.find_element_by_xpath(".//*[@class='table']/tbody/tr/td[4]").click()
                #     time.sleep(8)
                #     driver.switch_to.window(driver.window_handles[1])  # 切换当前页面标签
                #     driver = changepage(driver)  # 切换到风险提示页面
                #     companyannualreportinfo = driver.page_source
                #     company_annualreportinfo_initiatorscontributions(driver, company, companyannualreportinfo)  # 发起人及投资
                # 经营状况
                # operational_status_num = driver.find_element_by_xpath(".//*[@class='table']/tbody/tr/td[5]/span[2]").text
                # if operational_status_num == 0:
                #     print("该公司没有经营情况")
                # else:
                #     print("该公司有经营情况")
                #     driver.find_element_by_xpath(".//*[@class='table']/tbody/tr/td[5]").click()
                #     driver.switch_to.window(driver.window_handles[1])  # 切换当前页面标签
                #     driver = changepage(driver)  # 切换到风险提示页面
                #     operational_status = driver.page_source
                #     company_operationalstatus_administrativeLicense(driver, company, operational_status)  # 行政许可
                #     company_operationalstatus_qualitysupervisi(driver, company, operational_status)  # 质量监督检查

            driver.quit()
if __name__ == '__main__':
    filename = './companylist.xlsx'
    workbook = load_workbook(filename)
    sheet = workbook.active
    rows = sheet.max_row
    # for j in range(1, int(rows/5)+1):
    for i in range(1, 2):
        cname = sheet.cell(i, 1).value
        threading.Thread(target=company_info, args=(cname,)).start()
        print(cname)
        # time.sleep(20)
    # for k in range(1, 3):
    #     sheet.delete_rows(1)
    #     workbook.save(filename)
# cur.close()
# # conn.close()